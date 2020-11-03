import time
import random
import requests
import openpyxl

import signaturehelper

import pandas as pd
from pandas import DataFrame

def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(round(time.time() * 1000))
    signature = signaturehelper.Signature.generate(timestamp, method, uri, SECRET_KEY)
    return {'Content-Type': 'application/json; charset=UTF-8', 'X-Timestamp': timestamp, 'X-API-KEY': API_KEY, 'X-Customer': str(CUSTOMER_ID), 'X-Signature': signature}


BASE_URL = 'https://api.naver.com'
API_KEY = '0100000000adc996fd9c6660f2496ec0f64a1ce8c5688874b0d0f5074d8f98b0fadd4793b9'
SECRET_KEY = 'AQAAAACtyZb9nGZg8kluwPZKHOjFpb6b54doFnHnBVokUWhNFw=='
CUSTOMER_ID = '1158940'




wb = openpyxl.load_workbook('keyword.xlsx')
sheet1 = wb['Sheet1']

#'목적_기계/부품 등 정밀_PC' 
#'방식_밀링_PC'
#'재료_플라스틱_PC'
#'재료_금속_PC'
#'재료_기타재료_PC'
#'방식_CNC/MCT/머시닝_PC'
#'방식_축_PC'
#'재료_아크릴_PC'
#'목적_시제품/목업/졸업작품 등 일반_PC'
#'방식_금형_PC'
#'재료_알루미늄_PC'
#'방식_선반_PC'
group1 = []# grp-a001-01-000000017679946
group2 = [] #grp-a001-01-000000017679947	
group3 = []# grp-a001-01-000000017679952
group4 = []# grp-a001-01-000000017679958
group5 = []# grp-a001-01-000000017679961
group6 = []# grp-a001-01-000000017679962
group7 = []# grp-a001-01-000000017679966
group8 = []# grp-a001-01-000000017679969
group9 = []# grp-a001-01-000000017679972
group10 = []# grp-a001-01-000000017679976
group11 = []# grp-a001-01-000000017679980
group12 = []# grp-a001-01-000000017679982





def keyword():
   
    for i in range(0,992) :
        keyword = sheet1.cell(row = i + 2, column = 1).value
        group = sheet1.cell(row = i + 2, column = 2).value
        #  if('목적_기계/부품 등 정밀_PC' in group) :
        #     group1.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group1:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679946'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]

        # if('방식_밀링_PC' in group) :
        #     group2.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group2:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679947'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group2)


        # if('재료_플라스틱_PC' in group) :
        #     group3.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group3:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679952'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         print(group3)
        #         created_adkeyword = r.json()[0]

        # if('재료_금속_PC' in group) :
        #     group4.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group4:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679958'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group4)

        # if('재료_기타재료_PC' in group) :
        #     group5.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group5:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679961'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(i)


        # if('방식_CNC/MCT/머시닝_PC' in group) :
        #     group6.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group6:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679962'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group6)


        # if('방식_축_PC' in group) :
        #     group7.append(keyword)
 
        #     uri = '/ncc/keywords'
        #     method = 'POST'
 
        #     for i in group7:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679966'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group7)
 
        # if('재료_아크릴_PC' in group) :
        #     group8.append(keyword)
 
        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group8:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679969'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group8)


        # if('목적_시제품/목업/졸업작품 등 일반_PC' in group) :
        #     group9.append(keyword)
 
        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group9:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679972'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group9)


        # if('방식_금형_PC' in group) :
        #     group10.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group10:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679976'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group10)


        # if('재료_알루미늄_PC' in group) :
        #     group11.append(keyword)

        #     uri = '/ncc/keywords'
        #     method = 'POST'

        #     for i in group11:    
        #         r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679980'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
        #         #print("response status_code = {}".format(r.status_code))
        #         #print("response body = {}".format(r.json()))
        #         created_adkeyword = r.json()[0]
        #         print(group11)


        if('방식_선반_PC' in group) :
            group12.append(keyword)

            uri = '/ncc/keywords'
            method = 'POST'

            for i in group12:    
                r = requests.post(BASE_URL + uri, params={'nccAdgroupId': 'grp-a001-01-000000017679982'}, json=[{'keyword': i}], headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
                #print("response status_code = {}".format(r.status_code))
                #print("response body = {}".format(r.json()))
                created_adkeyword = r.json()[0]
                print(group12)



keyword()
# print(group1)

# for n in range(5):
#     data = sheet1.cell(row=n+2, column=3).value
#     if ('_1' in data) | ('(1)' in data):
#         sheet1.cell(row=n+2, column=2).value = '광고1'
#     elif ('_2' in data) | ('(2)' in data):
#         sheet1.cell(row=n+2, column=2).value = '광고2'
#     elif ('_3' in data) | ('(3)' in data):
#         sheet1.cell(row=n+2, column=2).value = '광고3'
#     else :
#         sheet1.cell(row=n+2, column=2).value = ''
    

# wb.save('keyword.xlsx')