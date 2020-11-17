import openpyxl

data = []

wb = openpyxl.load_workbook('광고 보고서 (2).xlsx')
sheet1 = wb['Sheet0']

ad_group = sheet1.cell(row=3, column=37).value

for n in range(390):
    ad_group_name = sheet1.cell(row=n+4, column=37).value
    data = sheet1.cell(row=n+4, column=30).value
    if '정석' in ad_group_name:
        print(n)
        # if ('_1' in data) | ('(1)' in data):
        #     sheet1.cell(row=n+2, column=2).value = '광고1'
        # elif ('_2' in data) | ('(2)' in data):
        #     sheet1.cell(row=n+2, column=2).value = '광고2'
        # elif ('_3' in data) | ('(3)' in data):
        #     sheet1.cell(row=n+2, column=2).value = '광고3'
        # elif ('_1' not in data) & ('_2' not in data) & ('_3' not in data) & ('--' not in data):
        #     sheet1.cell(row=n+2, column=2).value = ''
    elif 'B급' in ad_group_name:
        print(n)
    #     if '01' in data :
    #         sheet1.cell(row=n+2, column=2).value = '외근없는 재택 제조 받아보는 비교 견적서'
    #     elif '03' in data :
    #         sheet1.cell(row=n+2, column=2).value = '가공업체 찾는가? 비교견적 받았는가? 견적서 3개 받기'
    #     elif '04_336' in data :
    #         sheet1.cell(row=n+2, column=2).value = 'CNC 가공업체 찾기 힘들죠? 수수료 없이 비교견적받고 퇴근하세요'
    #     elif '04_ㅅ' in data :
    #         sheet1.cell(row=n+2, column=2).value = '시제품 비교견적 무료'
    #     elif '04_ㅇㅗ' in data :
    #         sheet1.cell(row=n+2, column=2).value = '부품가공 비교견적 무료'
    #     elif '04_ㅇㅣㄴ' in data :
    #         sheet1.cell(row=n+2, column=2).value = 'CNC 가공업체 찾기 힘들죠? 수수료 없이 비교견적받고 퇴근하세요'
    #     elif '05' in data :
    #         sheet1.cell(row=n+2, column=2).value = '가공업체? 찾지말고 받으세요 대기업 벤더 무료로 얻기'
    #     elif '07_ㅂㅏ' in data :
    #         sheet1.cell(row=n+2, column=2).value = '가공 업체 비교견적 수수료 없이'
    #     elif '07_ㅈㅓ' in data :
    #         sheet1.cell(row=n+2, column=2).value = '가공 업체 비교견적 마진없이 하루만에'
    #     elif '07_ㅋㅡ' in data :
    #         sheet1.cell(row=n+2, column=2).value = '가공 업체 비교견적 마진없이 무료'
    else :
        continue

wb.save('test.xlsx')

#37row = 광고그룹
#30row = 광고이름