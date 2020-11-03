import openpyxl

data = []

wb = openpyxl.load_workbook('test.xlsx')
sheet1 = wb['Sheet1']

for n in range(5):
    data = sheet1.cell(row=n+2, column=3).value
    if ('_1' in data) | ('(1)' in data):
        sheet1.cell(row=n+2, column=2).value = '광고1'
    elif ('_2' in data) | ('(2)' in data):
        sheet1.cell(row=n+2, column=2).value = '광고2'
    elif ('_3' in data) | ('(3)' in data):
        sheet1.cell(row=n+2, column=2).value = '광고3'
    else :
        sheet1.cell(row=n+2, column=2).value = ''
    

wb.save('test.xlsx')